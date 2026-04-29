"""
Module Export - Xuất báo cáo Điểm Danh sang Excel.
Sử dụng openpyxl để tạo file Excel với định dạng chuẩn.
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import logging

logger = logging.getLogger(__name__)

class ExcelExporter:
    """
    Lớp xuất báo cáo Điểm Danh sang Excel.
    """
    
    @staticmethod
    def export_attendance_to_excel(session_data, attendance_data, output_path):
        """
        Xuất kết quả 1 phiên điểm danh sang file Excel.
        
        Args:
            session_data (dict): Thông tin phiên từ DB
                {
                    'id': int,
                    'ngay_diem_danh': str (YYYY-MM-DD),
                    'tong_so_sv': int,
                    'co_mat': int,
                    'vang_mat': int
                }
            attendance_data (list): Danh sách dict, mỗi dict:
                {
                    'mssv': str,
                    'ho_ten': str,
                    'trang_thai': str ('CÓ MẶT' hoặc 'VẮNG MẶT'),
                    'similarity_score': float (optional)
                }
            output_path (str): Đường dẫn save file Excel
        
        Returns:
            bool: True nếu thành công
        """
        try:
            # Tạo workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Điểm Danh"
            
            # ===== HEADER STYLE =====
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # ===== BORDER =====
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # ===== ROW 1-3: THÔNG TIN CHUNG =====
            ws['A1'] = "BÁO CÁO ĐIỂM DANH"
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:F1')
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
            
            ws['A2'] = f"Ngày: {session_data.get('ngay_diem_danh', 'N/A')}"
            ws.merge_cells('A2:F2')
            ws['A2'].font = Font(size=11)
            
            # ===== THỐNG KÊ =====
            ws['A4'] = "Tổng Số SV:"
            ws['B4'] = session_data.get('tong_so_sv', 0)
            ws['B4'].font = Font(bold=True, size=11)
            
            ws['D4'] = "Có Mặt:"
            ws['E4'] = session_data.get('co_mat', 0)
            ws['E4'].font = Font(bold=True, color="008000", size=11)
            
            ws['A5'] = "Vắng Mặt:"
            ws['B5'] = session_data.get('vang_mat', 0)
            ws['B5'].font = Font(bold=True, color="FF0000", size=11)
            
            ws['D5'] = "Tỷ Lệ Có Mặt:"
            total = session_data.get('tong_so_sv', 1)
            co_mat = session_data.get('co_mat', 0)
            rate = f"{(co_mat / total * 100):.1f}%" if total > 0 else "0%"
            ws['E5'] = rate
            ws['E5'].font = Font(bold=True, size=11)
            
            # ===== HEADER CỘT =====
            headers = ["STT", "MSSV", "Họ Tên", "Trạng Thái", "Độ Tương Đồng", "Ghi Chú"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=7, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # ===== DỮ LIỆU =====
            center_alignment = Alignment(horizontal="center", vertical="center")
            left_alignment = Alignment(horizontal="left", vertical="center")
            
            for row_idx, record in enumerate(attendance_data, start=8):
                # STT
                ws.cell(row=row_idx, column=1).value = row_idx - 7
                
                # MSSV
                ws.cell(row=row_idx, column=2).value = record.get('mssv', '')
                
                # Họ Tên
                ws.cell(row=row_idx, column=3).value = record.get('ho_ten', '')
                
                # Trạng Thái
                trang_thai = record.get('trang_thai', 'VẮNG MẶT')
                cell = ws.cell(row=row_idx, column=4)
                cell.value = trang_thai
                
                # Màu theo trạng thái
                if trang_thai == 'CÓ MẶT':
                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                    cell.font = Font(color="155724")
                else:
                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                    cell.font = Font(color="721C24")
                
                cell.alignment = center_alignment
                
                # Độ Tương Đồng
                similarity = record.get('similarity_score', '')
                if similarity != '':
                    ws.cell(row=row_idx, column=5).value = f"{similarity:.4f}"
                else:
                    ws.cell(row=row_idx, column=5).value = '-'
                
                # Ghi Chú
                ws.cell(row=row_idx, column=6).value = ''
                
                # Apply border tất cả cells
                for col in range(1, 7):
                    ws.cell(row=row_idx, column=col).border = thin_border
                    if col == 1 or col == 4 or col == 5:
                        ws.cell(row=row_idx, column=col).alignment = center_alignment
                    else:
                        ws.cell(row=row_idx, column=col).alignment = left_alignment
            
            # ===== ADJUST COLUMN WIDTHS =====
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 20
            
            # ===== SAVE FILE =====
            os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
            wb.save(output_path)
            
            logger.info(f"✓ Xuất Excel thành công: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Lỗi export_attendance_to_excel: {e}")
            return False

    @staticmethod
    def export_month_report(month, year, sessions_data, output_path):
        """
        Xuất báo cáo tháng - tóm tắt tất cả phiên điểm danh trong 1 tháng.
        
        Args:
            month (int): Tháng (1-12)
            year (int): Năm
            sessions_data (list): Danh sách phiên trong tháng đó
            output_path (str): Đường dẫn save
        
        Returns:
            bool: True nếu thành công
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = f"Tháng {month}/{year}"
            
            # Header
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            ws['A1'] = f"BÁO CÁO THÁNG {month}/{year}"
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:G1')
            ws['A1'].alignment = Alignment(horizontal="center")
            
            # Header cột
            headers = ["Ngày", "Tổng SV", "Có Mặt", "Vắng Mặt", "Tỷ Lệ %", "Ghi Chú", "Chi Tiết"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Dữ liệu
            for row_idx, session in enumerate(sessions_data, start=4):
                ws.cell(row=row_idx, column=1).value = session.get('ngay_diem_danh', '')
                ws.cell(row=row_idx, column=2).value = session.get('tong_so_sv', 0)
                ws.cell(row=row_idx, column=3).value = session.get('co_mat', 0)
                ws.cell(row=row_idx, column=4).value = session.get('vang_mat', 0)
                
                total = session.get('tong_so_sv', 1)
                co_mat = session.get('co_mat', 0)
                rate = f"{(co_mat / total * 100):.1f}%" if total > 0 else "0%"
                ws.cell(row=row_idx, column=5).value = rate
                
                ws.cell(row=row_idx, column=6).value = ''
                ws.cell(row=row_idx, column=7).value = f"Session {session.get('id', '')}"
                
                # Border
                for col in range(1, 8):
                    ws.cell(row=row_idx, column=col).border = thin_border
            
            # Column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 15
            
            os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
            wb.save(output_path)
            
            logger.info(f"✓ Xuất báo cáo tháng: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Lỗi export_month_report: {e}")
            return False


def export_attendance_to_excel(database, session_id, output_path):
    """
    Convenience function để export kết quả từ database.
    
    Args:
        database: Database instance
        session_id (int): ID phiên điểm danh
        output_path (str): Đường dẫn save Excel
    
    Returns:
        bool: True nếu thành công
    """
    try:
        session = database.get_session_by_id(session_id)
        if not session:
            logger.error(f"Session {session_id} không tồn tại")
            return False
        
        attendance = database.get_attendance_by_session(session_id)
        
        return ExcelExporter.export_attendance_to_excel(session, attendance, output_path)
        
    except Exception as e:
        logger.error(f"Lỗi export_attendance_to_excel: {e}")
        return False
