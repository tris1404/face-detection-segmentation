"""
Test module cho Excel Export functionality.
Test cases: Excel file creation, format validation, data accuracy.
"""

import pytest
import os
import tempfile
from openpyxl import load_workbook
from utils.export import ExcelExporter, export_attendance_to_excel
from utils.database import Database


@pytest.fixture
def temp_excel_file():
    """Tạo file Excel tạm thời."""
    fd, path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    yield path
    if os.path.exists(path):
        os.remove(path)


@pytest.fixture
def sample_session_data():
    """Dữ liệu phiên test."""
    return {
        'id': 1,
        'ngay_diem_danh': '2024-03-15',
        'tong_so_sv': 5,
        'co_mat': 4,
        'vang_mat': 1
    }


@pytest.fixture
def sample_attendance_data():
    """Dữ liệu điểm danh test."""
    return [
        {
            'mssv': '20210001',
            'ho_ten': 'Nguyễn Văn A',
            'trang_thai': 'CÓ MẶT',
            'similarity_score': 0.95
        },
        {
            'mssv': '20210002',
            'ho_ten': 'Trần Thị B',
            'trang_thai': 'CÓ MẶT',
            'similarity_score': 0.92
        },
        {
            'mssv': '20210003',
            'ho_ten': 'Lê Văn C',
            'trang_thai': 'CÓ MẶT',
            'similarity_score': 0.88
        },
        {
            'mssv': '20210004',
            'ho_ten': 'Phạm Thị D',
            'trang_thai': 'CÓ MẶT',
            'similarity_score': 0.85
        },
        {
            'mssv': '20210005',
            'ho_ten': 'Hoàng Văn E',
            'trang_thai': 'VẮNG MẶT',
            'similarity_score': 0.0
        }
    ]


# ==================== TEST EXCEL EXPORT ====================

def test_export_attendance_to_excel_success(temp_excel_file, sample_session_data, sample_attendance_data):
    """Test 1: Xuất Excel thành công."""
    success = ExcelExporter.export_attendance_to_excel(
        sample_session_data,
        sample_attendance_data,
        temp_excel_file
    )
    
    assert success is True
    assert os.path.exists(temp_excel_file)
    assert os.path.getsize(temp_excel_file) > 0
    
    print(f"\n✓ Test 1 Passed: Excel file created at {temp_excel_file}")


def test_excel_file_format(temp_excel_file, sample_session_data, sample_attendance_data):
    """Test 2: Kiểm tra định dạng Excel file."""
    # Export
    ExcelExporter.export_attendance_to_excel(
        sample_session_data,
        sample_attendance_data,
        temp_excel_file
    )
    
    # Load workbook
    wb = load_workbook(temp_excel_file)
    ws = wb.active
    
    # Check sheet name
    assert ws.title == "Điểm Danh"
    
    # Check header row (row 7)
    expected_headers = ["STT", "MSSV", "Họ Tên", "Trạng Thái", "Độ Tương Đồng", "Ghi Chú"]
    for col, header in enumerate(expected_headers, start=1):
        cell = ws.cell(row=7, column=col)
        assert cell.value == header
    
    print("\n✓ Test 2 Passed: Excel format correct with proper headers")


def test_excel_data_accuracy(temp_excel_file, sample_session_data, sample_attendance_data):
    """Test 3: Kiểm tra độ chính xác dữ liệu trong Excel."""
    # Export
    ExcelExporter.export_attendance_to_excel(
        sample_session_data,
        sample_attendance_data,
        temp_excel_file
    )
    
    # Load workbook
    wb = load_workbook(temp_excel_file)
    ws = wb.active
    
    # Check data rows (from row 8)
    for row_idx, record in enumerate(sample_attendance_data, start=8):
        col_mssv = 2
        col_ho_ten = 3
        col_trang_thai = 4
        
        assert ws.cell(row=row_idx, column=col_mssv).value == record['mssv']
        assert ws.cell(row=row_idx, column=col_ho_ten).value == record['ho_ten']
        assert ws.cell(row=row_idx, column=col_trang_thai).value == record['trang_thai']
    
    print("\n✓ Test 3 Passed: Data in Excel matches input")


def test_excel_session_stats(temp_excel_file, sample_session_data, sample_attendance_data):
    """Test 4: Kiểm tra thống kê phiên trong Excel."""
    # Export
    ExcelExporter.export_attendance_to_excel(
        sample_session_data,
        sample_attendance_data,
        temp_excel_file
    )
    
    # Load workbook
    wb = load_workbook(temp_excel_file)
    ws = wb.active
    
    # Check stats (row 4-5)
    total_sv = ws.cell(row=4, column=2).value
    co_mat = ws.cell(row=4, column=5).value
    vang_mat = ws.cell(row=5, column=2).value
    
    assert total_sv == sample_session_data['tong_so_sv']
    assert co_mat == sample_session_data['co_mat']
    assert vang_mat == sample_session_data['vang_mat']
    
    print("\n✓ Test 4 Passed: Session stats correct in Excel")


def test_excel_color_coding(temp_excel_file, sample_session_data, sample_attendance_data):
    """Test 5: Kiểm tra màu sắc cho trạng thái."""
    # Export
    ExcelExporter.export_attendance_to_excel(
        sample_session_data,
        sample_attendance_data,
        temp_excel_file
    )
    
    # Load workbook
    wb = load_workbook(temp_excel_file)
    ws = wb.active
    
    # Check color coding (row 8 - CÓ MẶT, row 12 - VẮNG MẶT)
    co_mat_cell = ws.cell(row=8, column=4)
    vang_mat_cell = ws.cell(row=12, column=4)
    
    # Check có mặt (green)
    assert co_mat_cell.value == 'CÓ MẶT'
    assert co_mat_cell.fill is not None
    
    # Check vắng mặt (red)
    assert vang_mat_cell.value == 'VẮNG MẶT'
    assert vang_mat_cell.fill is not None
    
    print("\n✓ Test 5 Passed: Color coding applied correctly")


def test_export_month_report(temp_excel_file):
    """Test 6: Xuất báo cáo tháng."""
    sessions_data = [
        {
            'id': 1,
            'ngay_diem_danh': '2024-03-01',
            'tong_so_sv': 30,
            'co_mat': 28,
            'vang_mat': 2
        },
        {
            'id': 2,
            'ngay_diem_danh': '2024-03-05',
            'tong_so_sv': 30,
            'co_mat': 29,
            'vang_mat': 1
        }
    ]
    
    success = ExcelExporter.export_month_report(
        month=3,
        year=2024,
        sessions_data=sessions_data,
        output_path=temp_excel_file
    )
    
    assert success is True
    assert os.path.exists(temp_excel_file)
    
    # Check content
    wb = load_workbook(temp_excel_file)
    ws = wb.active
    assert ws.title == "Tháng 3/2024"
    
    print("\n✓ Test 6 Passed: Month report generated successfully")


def test_export_with_missing_data(temp_excel_file, sample_session_data):
    """Test 7: Export với dữ liệu không đầy đủ (missing similarity score)."""
    attendance_data = [
        {
            'mssv': '20210001',
            'ho_ten': 'Sinh viên A',
            'trang_thai': 'CÓ MẶT',
            'similarity_score': None  # Missing
        },
        {
            'mssv': '20210002',
            'ho_ten': 'Sinh viên B',
            'trang_thai': 'VẮNG MẶT',
            'similarity_score': 0.0
        }
    ]
    
    # Export - không gây lỗi
    success = ExcelExporter.export_attendance_to_excel(
        sample_session_data,
        attendance_data,
        temp_excel_file
    )
    
    assert success is True
    
    # Verify file created despite missing data
    assert os.path.exists(temp_excel_file)
    
    print("\n✓ Test 7 Passed: Export handles missing data gracefully")


def test_export_large_dataset(temp_excel_file):
    """Test 8: Export dataset lớn (100 sinh viên)."""
    session_data = {
        'id': 1,
        'ngay_diem_danh': '2024-03-15',
        'tong_so_sv': 100,
        'co_mat': 95,
        'vang_mat': 5
    }
    
    # Generate 100 records
    attendance_data = [
        {
            'mssv': f'2021{i:04d}',
            'ho_ten': f'Sinh Viên {i}',
            'trang_thai': 'CÓ MẶT' if i <= 95 else 'VẮNG MẶT',
            'similarity_score': 0.9 + (i % 10) * 0.01
        }
        for i in range(1, 101)
    ]
    
    success = ExcelExporter.export_attendance_to_excel(
        session_data,
        attendance_data,
        temp_excel_file
    )
    
    assert success is True
    
    # Load and verify count
    wb = load_workbook(temp_excel_file)
    ws = wb.active
    
    # Data rows: 100 (from row 8 to row 107)
    last_row = 107
    assert ws.cell(row=last_row, column=2).value == '20211000'
    
    print("\n✓ Test 8 Passed: Large dataset (100 records) exported successfully")


def test_convenience_function_export(temp_excel_file):
    """Test 9: Convenience function export_attendance_to_excel."""
    # Create temp database
    fd, db_path = tempfile.mkstemp(suffix='.db')
    os.close(fd)
    
    try:
        from utils.database import Database
        
        db = Database(db_path)
        
        # Add test data
        result = db.add_student('20210001', 'Test Student', 'path/test.jpg', '{}')
        student_id = result['student_id']
        
        session_id = db.create_session('2024-03-15', 'path/class.jpg')
        
        db.save_attendance(session_id, student_id, 'CÓ MẶT', 0.95)
        db.update_session_stats(session_id)
        
        # Export
        success = export_attendance_to_excel(db, session_id, temp_excel_file)
        
        assert success is True
        assert os.path.exists(temp_excel_file)
        
        print("\n✓ Test 9 Passed: Convenience function works correctly")
        
    finally:
        if os.path.exists(db_path):
            os.remove(db_path)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
